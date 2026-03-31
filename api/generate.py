from http.server import BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import json, io, base64

# ── LOGO (base64 embedded) ──────────────────────────────────────────────────
LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAhcAAAAnCAIAAAAzXv2bAAAKHmlDQ1BJQ0MgUHJvZmlsZQAAeJy1Vnk8lGsbft73nX2xzZDd2LdGljDIvpPITpsxMxjLYMyg0iapcCJJthI5FTp0WpDTIi3ajtKmos7IEarT0SKVyvcOf+j7fefP812/3/O813v97vt+7ud+/3gvAMhjAAWMrhSBSBjs7caIjIpm4B8DBKgBRaAHtNicjDTwv4Dm6ceHc2/3mNLd+JPjs9Z3YS3Zbl/+vLHVjvoPuT9CjsvL4KDlPFC+NhY9HOVdKKfHhga7o/w+AAQKN4XLBYAoQfUd8bMxpARpTPwPMcniFD6q50j1FB47A+UlKNeLTUoTofyUVBfO5V6b5T/kingctB5pENUpmWIeehZJOpftWSJpLll6fzonTSjleSi35SSw0RjyWZQvnOt/FloZ0gH6errbWNjZ2DAtmRaM2GQ2J4mRwWEnS6v+25B+qzmmdxAAWbS3ttscsTBzTsNINywgAVlABypAE+gCI8AElsAWOAAX4An8QCAIBVFgNeCABJAChCAL5IAtIB8UghKwF1SBWtAAGkELOAHawVlwEVwFN8Ed8AAMAAkYAa/ABPgIpiEIwkNUiAapQFqQPmQKWUIsyAnyhJZCwVAUFAPFQwJIDOVAW6FCqBSqguqgRuhX6Ax0EboO9UGPoSFoHHoHfYERmALTYQ3YAF4Es2BX2B8OhVfB8XA6vA7Og3fBFXA9fAxugy/CN+EHsAR+BU8iACEjSog2wkRYiDsSiEQjcYgQ2YgUIOVIPdKCdCI9yD1EgrxGPmNwGBqGgWFiHDA+mDAMB5OO2YgpwlRhjmLaMJcx9zBDmAnMdywVq441xdpjfbGR2HhsFjYfW449jD2NvYJ9gB3BfsThcEo4Q5wtzgcXhUvErccV4fbjWnFduD7cMG4Sj8er4E3xjvhAPBsvwufjK/HH8Bfwd/Ej+E8EMkGLYEnwIkQTBIRcQjmhiXCecJcwSpgmyhH1ifbEQCKXuJZYTGwgdhJvE0eI0yR5kiHJkRRKSiRtIVWQWkhXSIOk92QyWYdsR15O5pM3kyvIx8nXyEPkzxQFignFnbKSIqbsohyhdFEeU95TqVQDqgs1miqi7qI2Ui9Rn1E/ydBkzGR8Zbgym2SqZdpk7sq8kSXK6su6yq6WXSdbLntS9rbsazminIGcuxxbbqNctdwZuX65SXmavIV8oHyKfJF8k/x1+TEFvIKBgqcCVyFP4ZDCJYVhGkLTpbnTOLSttAbaFdoIHUc3pPvSE+mF9F/ovfQJRQXFxYrhitmK1YrnFCVKiJKBkq9SslKx0gmlh0pfFmgscF3AW7BzQcuCuwumlNWUXZR5ygXKrcoPlL+oMFQ8VZJUdqu0qzxVxaiaqC5XzVI9oHpF9bUaXc1BjaNWoHZC7Yk6rG6iHqy+Xv2Q+i31SQ1NDW+NNI1KjUsarzWVNF00EzXLNM9rjmvRtJy0+FplWhe0XjIUGa6MZEYF4zJjQltd20dbrF2n3as9rWOoE6aTq9Oq81SXpMvSjdMt0+3WndDT0gvQy9Fr1nuiT9Rn6Sfo79Pv0Z8yMDSIMNhu0G4wZqhs6Gu4zrDZcNCIauRslG5Ub3TfGGfMMk4y3m98xwQ2sTZJMKk2uW0Km9qY8k33m/YtxC60WyhYWL+wn0lhujIzmc3MITMls6VmuWbtZm8W6S2KXrR7Uc+i7+bW5snmDeYDFgoWfha5Fp0W7yxNLDmW1Zb3rahWXlabrDqs3i42XcxbfGDxI2uadYD1dutu6282tjZCmxabcVs92xjbGtt+Fp0VxCpiXbPD2rnZbbI7a/fZ3sZeZH/C/m8HpkOSQ5PD2BLDJbwlDUuGHXUc2Y51jhInhlOM00EnibO2M9u53vm5i64L1+Wwy6irsWui6zHXN27mbkK3025T7vbuG9y7PBAPb48Cj15PBc8wzyrPZ146XvFezV4T3tbe6727fLA+/j67ffp9NXw5vo2+E362fhv8LvtT/EP8q/yfLzVZKlzaGQAH+AXsCRhcpr9MsKw9EAT6Bu4JfBpkGJQe9Nty3PKg5dXLXwRbBOcE94TQQtaENIV8DHULLQ4dCDMKE4d1h8uGrwxvDJ+K8IgojZBELorcEHkzSjWKH9URjY8Ojz4cPbnCc8XeFSMrrVfmr3y4ynBV9qrrq1VXJ68+t0Z2DXvNyRhsTERMU8xXdiC7nj0Z6xtbEzvBcefs47ziunDLuOM8R14pbzTOMa40bizeMX5P/HiCc0J5wmu+O7+K/zbRJ7E2cSopMOlI0kxyRHJrCiElJuWMQEGQJLicqpmandqXZpqWnyZJt0/fmz4h9BcezoAyVmV0iOjoD+aW2Ei8TTyU6ZRZnfkpKzzrZLZ8tiD71lqTtTvXjq7zWvfzesx6zvruHO2cLTlDG1w31G2ENsZu7N6kuylv08hm781Ht5C2JG35Pdc8tzT3w9aIrZ15Gnmb84a3eW9rzpfJF+b3b3fYXrsDs4O/o3en1c7Knd8LuAU3Cs0Lywu/FnGKbvxk8VPFTzO74nb1FtsUHyjBlQhKHu523n20VL50XenwnoA9bWWMsoKyD3vX7L1evri8dh9pn3ifpGJpRUelXmVJ5deqhKoH1W7VrTXqNTtrpvZz99894HKgpVajtrD2y0H+wUd13nVt9Qb15YdwhzIPvWgIb+j5mfVz42HVw4WHvx0RHJEcDT56udG2sbFJvam4GW4WN48fW3nszi8ev3S0MFvqWpVaC4+D4+LjL3+N+fXhCf8T3SdZJ1tO6Z+qOU07XdAGta1tm2hPaJd0RHX0nfE7093p0Hn6N7PfjpzVPlt9TvFc8XnS+bzzMxfWXZjsSut6fTH+4nD3mu6BS5GX7l9efrn3iv+Va1e9rl7qce25cM3x2tnr9tfP3GDdaL9pc7PtlvWt079b/36616a37bbt7Y47dnc6+5b0nb/rfPfiPY97V+/73r/5YNmDvodhDx/1r+yXPOI+Gnuc/Pjtk8wn0wObB7GDBU/lnpY/U39W/4fxH60SG8m5IY+hW89Dng8Mc4Zf/Znx59eRvBfUF+WjWqONY5ZjZ8e9xu+8XPFy5FXaq+nX+X/J/1XzxujNqb9d/r41ETkx8lb4duZd0XuV90c+LP7QPRk0+exjysfpqYJPKp+OfmZ97vkS8WV0Ousr/mvFN+Nvnd/9vw/OpMzM/OBNzFBbwpj3JR68OLY4WcSQGhb31ORUsZARksbm8BhMhtTE/N98SmwlAO3bAFB+Mq+hCJp7zPm2WUDgnwHP5yFK6LJCpYZ5LbUeANYkqpdk8ONnNffgUMYPc2AG8+J4Qp4AvWo4n5fFF8Sj9xdw+SJ+qoDBFzD+a0z/yuV/wHyf855ZxMsWzfaZmrZWyI9PEDF8BSKeUMCWdsROnv06QmmPGalCEV+cspBhaW5uB0BGnJXlbCmIgnpn7B8zM+8NAMCXAfCteGZmum5m5hs6C2QAgC7xfwAKP9n2U7+jGwAAROxJREFUeJztXXmczeX+f77b+Z59G7NaJkuULhFFlqRsFVLWRIREiRShspQlbjKKItpsWUr8krTcqLh1E5Fd0hXGLGbmzNnP+a6/P97O8zuNOYc5Tfe2/D4vL68zM9/zfJ/l83z2hZk5dfrOnTsVRdE0LRKJWCwWVVVZllUUhVQGDMOoqipJkt1ul2VZkiSXyxUKhRiB9/l8NptNlmVBEAKBQKNGjVavWUOYSochhJBIJCIIAsdxmqaxLEsIUVWVYzlCyP9s3jxjxgyXy6VpmsFg2LBhg9Vu43meDs7HIOHohBBCdF1nmAsz+PTTT8eMGZPlrqFoaoMGDVasXk0IwfQkSTIYDMmHumhoQgg5+eOP999/P8uygUCgXr1669evZzi20sexwwaDYdwjYw8fPhwOBgkhRNMTLcFoNOq6znGc0Wi02+25ubktWrRo0qRJvUYNia4ThsH/qqKoqmoQRR0TIoSuNzmoqspxHCFEikQJIQLHdevWLRwMWSyWsnLP8OHDR4wcqes6J/CapsmyLIpi1fYnKcSfy+HDh/v27ZvhTsNvGIbRNI3jOEVRbrjhhufzFlTXSwOBQPPmzffu3Wu32wkhFOsqBVmWu3fv/uijj952222apkmSBJxP4b3xi9U0rVOnTvPmzbv++utJ7BRUVY1Go/fdd195eXk4HE40js1m83g8TqezrKwM9zQSiSTCW0EQ/H6/y+WqUaPGsmXLMPNoNIpz/PrrrydMmGAwGEKhkNFoTHTZk4DFYgmHwyzLDho0qF+/fhaLBZhcYVcVRfnuu+9GjhxpsVgqHQc7o+s6IcRgMNjt9rS0NLvdfsMNN+Tm5l5//fUGgwEbGAgErFYrIQQf6I+hUMhsNlcY9rnnnvP5fLNmzQKSXwyKovA8X1pa6nQ6GYZhWVZVVV3XeZ7HbmA50Wh03759u3bt2rNnT1lZWWFhYVlZma7rTqczMzPT6XQ2bNiwffv2zZs3z8zM5DhOkiSO4/ABR0M/lJaWpqWlJdnSEydODBgwwOVylZSUpKWlhcNhk8kUiUQqfTgtLQ1zaNCgQatWrZo1a0YIwYkIgoAjwL55vV6Hw4Fveb1ek8lkMBjoMcVv5pEjR7Zs2bJv377S0tKSkpJIJAJsTEtLa9SoUdOmTXNzc++55x46cmlpqdVqFUWR93jLzxUW+P1+7KPRaJQkief5RMRI13VFUWRZPpN/NisrS5blYz8cz87O9vr9PM/7/f7y8vJQKNSgQQOPxyNLkiAmpM5Go5GOiWlxHEc0PRQKrV+/vqioqKioSBTFtLS0xYsXPz1tKiFE0zRCCDDyMsllBfB4PKqu+Xw+ouuqqoJKAomrBLqmaZpWWlqan58Pppudna1pGpeAi1Bu4fP5zp07Fw2HCSFGgyirlV9gsFjc+Wg0+vmXX6xdv85sNhvN5v79+w8ePDg9I8Pn9dodDlVVqzp5QghWzfO8qqoms/mddeuKi4uD/kAoFDJZzO+8887gIUNwQMDLFF5RJThz5gzLskADsE9ZlrOzs3/r9yYCQRAKCwtnz57drFmz7OxsVVXtdrvP56sqI0mEWrjG4JeEELPZPGrUKLPZnIixSZIkiiIEuLfeestut99xxx2UdidagiAIkUjEbrdLkoSZRKNRUGSTyfTUU09ZLBZBEBRFqSoWGQwGlmWDweCVV17pcDii0SghJAlXfuGFFxLtA4g45Aafz1dcXOzxeHbs2JGfn19cXDxgwICxY8dqmma1WkEQrVarJElWq1VRFF3XL2YhlwM8zxcWFmZlZWmaBrkZBwGepCjK/v37N27c+NFHHwmC0KJFi969e7vd7pycnLS0NJZly8vLi4qKPB7PwYMH8/LyCgoK6tWrd++99w4YMICuC7uKKxwKhdLS0oqLizMyMiqdjyzL0WhUkqSJEydijT6fD9S40ufLy8tBSXbv3r106VKWZfv27TtkyJDs7Ozy8nKn00kI0XXd7/fHsxCHw6FpGuW7gUAATHrv3r0zZswoKCjo2LFjp06dGjdubDabTSaT1WoFrz106NDu3bu//vrryZMn9+7de8CAAQ0bNkxLS8MaL0j0mqbxPM9xHMuyDMNAHkx0AKIomkwm0FCGYTIyMoLBoNPp9Pl8qqpardb69ev/+OOP0CQSDQJ6AYaEV+P3kiQdO3bs66+/1nXdYrFomubxeD799NOHHxkD8TxesksC9LEKz/t8Pp0hoVAIeM8J/AXuVUUAsVNVVVEUbNflTEySpFAopCgKy7Jer5dY9ES6C1A5GAwSQkRRNBqNsix7PB4hFHrllVdWrlzZr1+/xx9/3O/zmc1mT1mZ0+2qsPDkIAgCEN1kMkUjkbfffhtnAe3w2LFjH3zwQd9+/QghoVCISjfVBZRbxIuiONwKH6r3pVV6PhwO5+bmPvHEEy+99JLL5QI5rq5p4P/4C9KpU6dLfhHbtXHjRpfL1bVr1yRnjaOkrzAYDNFo1GAwhMNh8KpwONyhQwdCCO57VVcBSoQxMTeqi1T6fOvWrS85Ji4URgAt2rNnz8qVK5s2bfroo48OHjzY4XBEIhGj0Yh38TwfiUR0XadzqLCQJCfu8/mysrJUVZVl2Wg0+nw+URQZhhFF8fjx48uXL//HP/7Rpk2bV199tUmTJmC08Rvlcrlyc3N1XW/btu3EiRPz8/N37NixZMmSl19++cknn7ztttuoBIbnYa5IxEIIIRAZy8rKOnTogO8Gg0Ge55PYAILBoK7r4Knffvvtu+++e8cdd3Tu3HnevHlgGLIsQ10D98KmqapqNpsjkQjP81ar9fz58/fff39hYeHjjz9+zz33QHilp4nnnU5ngwYNevfuzXHcqVOnli1bNnHixEaNGg0aNKh9+/aqqvI5tWo1u+46zBhCAch6IiU3PT397Nmz2PFAIOByuQwGQ1FRkcPhOHToUFlZWSAQUBQlPT199OjRYkzbqBR0XY9EIvGirq7rkiS9/vrruAMmk6mkpETX9fz8/A8//LBPnz66rsuyDI0MclzyCxCPWP/HgRhCRQ9CCMMwl7SMVQIMAwYsCAKsfDzPs4m5EUQe8A+O4wSDQdO05tc2442VYwnDMJBECgsLCwoKAsEgIYThOE3TjEZjeXn5ihUr9u/f/9xzz9WrV8/ldmu6RuLsA5dDFyD7E4bZtWvXwYMHbTabwWCQZZnjOEEQli9fftddd/EGwWg0ViMLqTA3es8hvpAYe2ZZNv6MfgtIzlQikYgoipMmTXrssccWLFjw9NNPp2zTq4CE+ECXhsWCOCqKkmjJDMPA+gSDCQbEtyp9HqJhvClSFEWqdguCAAGcGh6qymIxT9j6eJ6/pEE40fggZxgNU6JDBQKBpk2bLlq06NSpU8OHD9+xY8eCBQtq1qyJHYPBB1bu+PGpgJJ8RXa7HUYtjuPC4TCoLcuyS5YseeONNxo2bLh27dorr7ySMuMKJIKO73K5CCG5ubmDBw/u16/f6tWrp06dunr16lmzZlksloyMDJgrIftjzpXOB+ZubKwkSSDFF25ogudhklEURRCEtm3bNm/e/OzZszNmzOjcufPmzZtJzPsQCoUyMjIURbFYLKFQCAsJhUJut/vQoUOjRo26+uqr161bZ7VawUJAASRJAmUDvlEcTk9PnzNnztmzZ1euXPnoo48+//zzt956Kz/u0UdVRcEdVmIfGIYhiclQwO+32myRcBhE1FNW5nK7v/7qqxEjRoTDYYfDYbPZbr/99h49eiiyzBsSGkNAKKn9EdaVoqKi999/32KxBIPBzMzM8+fP22w2n8+3du3anj17Qo0FCwHhTown/3fe8WRLFEWN6JC4f6WoC+MejvySWAsLOMdxPM8HAgEDz9vt9h49egwZMbzy52PL5HkefHTXrl3/+te/tm/fDpYfCAS+/PLLYcOGvfbaa/Xr1yfsL/jHJRkJpASO42RJWrt2bTgc1nVd4Pi6deueKywwmUw//fTTp59+2r7DTUCvRNTq14MeA0rL9DioxreQuMt/SeA4ThRFh8OxYsWKu+66q379+kOHDqU0parvJZXxEkVRDAYDqAC2N8mpRaNR6KPg8XgsEUkisZWChcBaBeneYrHIsgweqaoqKBd1kl0+YGSDwRD/xWg0mmh/Eq0LS6CuCFAhRVECgUCNGjUIIZqmXXHFFZ988skjjzzy8MMPL1u2zOFwwByCv5pMJpA8chGvSn7W4H+YAwxrDz744L59+xYsWNC+fXvqfKUOGNAcmCLjnbKgAzjEkSNH9ujRY8KECQMHDly+fDkmabPZoPQkOS9JkmRZdrlckNhwykkmT7edqoAmk6lhw4arVq2aNm1az549V65cmZGRwbKs1WqF+weGYmilDoejpKRk3Lhx7dq1mzx5MgzIoihCM8PIkELoK2RZlmXZYrFIklSrVq0nn3zypptuuvHGG1VVZT3lHkmRGY7VGaIRXdU1fNAT/CsoLLDabTrRjWaTqmuhcMiV5j546OCoUaNCoRDkWaPROGnSJKPJxCfeCOitlIUAjw0Gw9q1ayFf6Lo+dOjQLl26GI1Gk8m0d+/e/fv3MwyDzQVFTrLL8XaD+P/BI3HHMAfoN0mGqnx8TaP4BIOPqqpaYuMy2DOeh+YuSVJ2dnaifeYEXhANosnI8pzOkJxaNfvfMyDvxYWbNm3q2LGjz+ezWq01atQoLy9//PHHodvGL5xcimJKkoQpff/99//617+sVmswGGzQoMG4cePC4bAoiqIoLlmyxGq1+v3+KoceJNm3OPYQP0PQOOgf+P+SR5zCq5P8WAEEQYAr8qqrrpo9e/asWbM+//zzX+kfisdDIABkPeChpmlQESoF7AnDMNFoFAIQHadSiEQikUhEVVXIBxXcHrhuoVAIP8IWXSWA/EQPCGwgyf4kGgd/hU0Yfg6e541GI1iILMuKokSj0VAotHDhwtzc3KlTp8ZjOGguZPYKl/2SLAQWOXxWVbVbt24FBQXbtm279tprCSHAfDhg4hcoiiLl4uArIErhcBi/tFqta9asufvuuwcOHHjgwAGz2YyNSi7yGo3GQCAAZoNxZFnG+In2LRqNUt+7oiiRSASEfsqUKa1bt37wwQepPSktLQ3sGbwfMvSKFSskSRo3bpzT6TSZTFiUwWAoLi7GtyizBOEyGAyI6cCsIpFI69at4c1iTSYTmA+GAEGkGsnFkJmZSQgJBAKEEFEUzWZzKBSaNm0aIaS8vJzneZZlN27c6HK7CSHRBAEGFAmwHVTv/ve//71q1SqM2aBBg7vuugv6Da7NmjVr/H4/RZQUSIweM7xCtY/nK1UdisRMYdg0LDz5w1TWNhgMJpMpGAyGQqFE+wwBB9ZYuElxog0bNVq+fPn06dM9Ho/X6y0uLj516hRclxczkiQAO5Wqqhs2bAiFQgaDwWazDRw48Pbbb69fv77f71dVdd++fSdPnoQfNYX9qRLIMYBQhg+JXIspw+XrIpFIJCMjo6SkhBBy2223DR069Nlnny0tLU3tpRU+67oObzmJUUAQZRITjy4GEGhRFHEcMF7DAlYpmEwmi8UC5YZanCA5QWrWNI36pV0uV6JxEkG8FQvUnyT1ricaBwZwjuMMBgPP85IkhcNhsA1CiCAIBoNBVVWbzSYIwowZM/bv379t2zZCiN/vx9Kgr1f1UKgU6/f7OY7r37+/2+1evXp1WloaxG3EVtE4K1xwWPAikUg0GoWki9FCoRB4fCgUEkXR5/ONHz9+0qRJffr0+emnn6hE7/V6k+MJJoPPIMWJ9g2KJs5RVVWQFJysruvPPPNMJBJ57733MFo0GoW2hBchOmDDhg1Tp05FAAtm6PP5gsEgnDegzOAF4G34LswSHMfBHgicZEVR1HUdPJnERIYkN41l2ZKSEuhoUHmmT5/+3XffFRQU5ObmSpK0ePFiBLQpcbt8MWDTgXYWi4XjuPLy8k2bNkWj0UAg4Ha7u3XrZrFab+rQwWq1qqrqcrnef/99j8cDixYWqVY9PImqIEycUz0FLGRYFoJhIBAIBAIIfUniF8FUIXNFo1GwsSSmduqtodYPTNvn9fKC8PCYMaNGjRJFMSsrq6SkZOXKlXR1lzl/xJUVFxdv3ryZ47hIJOJ2u3v37s2L4n333UelhHnz5omimEIk6OUDlQn4OKA/Vu9bLh/gfHK5XKAgTz/9NMMw06dPT/m9FSYAWxCOFXIDtT5VCrIsI+ySYRifzwdVEkbgJDI+pBCqtdDPcGoC/zFsonESARgh3nI5hr5E41itVojJ2AeYqiCeIrSEEALyzTCMy+UaPXp0Xl6epmk2mw0LxF+rei7gWCzL2my2KVOmRCKRvLw8sFWQToPBIIpiMBiEcEzD2GBrgUYCgg4LM/iN2WxGvDLP8/fee++4ceMGDx5M94qGS10Muq47HA4YNlRVxfNJ9j8cDuNWQgqJt6bYbDaGYcaMGbNgwQIMApcYorNkWQ4Gg0VFRaWlpV26dCGEeL1exPLZbDYakA1Gpcain6nB0GKxQO+xWq30pRfMLJScYV+ozediCPoDNdJqFBUUKpLMMuy6t9dufm+T2WhKz8zILzj3+MQJbdu3M1sthCEszxnNCe2AJHavvJ5youlyOKJEols2bmI41mg2CaJh7KPjCEMi0cjU6dNkVSGSInL87GeeZQmjKRc8Ikl0CCYWMErlI47joC0C51iOY2MewhSwkI4JcYm7lGLExdwwEEJhJKXMOxHQkGuq6+gM0XSNMGTy5Mk1a9aMhsOMrjOqNmPGDIrTbIzD0WOGKdzv9wPzJEkiOmEJs2rFSoSNsTz/8COPiFaLSrQhw4fZXE5QmW1bPig+V8CzPN0iVVURzXXJyVcKoGXYLupCpMq7FoN4w1e1AKwfNExe1xNm6gCi0SiccFzMlb18+fKDBw++9NJLeACMFiJzEi5LFwu2zXGcGvNSkNixUmsVSaqLmEwmyEAmk4muItHzeDvoHR0Z/8N2RG39MEUkGicRwHSBt8Sb+zEx6sYHAqsxp2Cl8+Q4rlJVnlK0+PiOwYMHy7K8a9cuoB/P86FQCJSUGo0JISB/SW4l6L4sy9u2bduxY8e8efNoFBMTF25jsVicTicYNhdHTCrMs1KTryRJkydPbtSo0ZgxY+IlwkoBojzGp6FZSXQRs9lMJ0mpPIlpFYqidO/e3Wq1bt++HciJ2A2z2QwtZ/fu3c2bN5ckyefzgbdd7MkH6sbTWEqLgDP/p9wkWlUisFit5R5PZlaWIAgHvv9+1qxZuq77fL6CgoKhQ4feeeedmC4MdsmpAAQlh9Op67pgNH7wwQeFhYX4bv/+/VmWhUexWbNmderUgQNt3759R44cEY1G2IJ+Uxn59wkWiwWoZjabu3XrJkmS0+nkef7o0aMUXfAktVlBXjMajRBSCCEGg0GWJE3T1q5dqygKYrJvvfVWEuN2wHsQrEWLFpGYAkcIgVAG3vkn3n9RFJFjQSlRvXr1Jk2atGjRop07d0KI9vl8EJmrUWf6f0gOiqL87W9/O3jwIOgs7HIp4CGobSQSWbhw4ahRo+rXrx9vkqmWeSKfZsGCBbt37/7oo4+qa+TkAK0CylOrVq12795NOZ8gCNRMV1xcnJmZGS/+Qq9NTW6rMhfRVNXpcgX8fk3TBg4ciGR1XdfbtWs3atSoOnXqkJiBiE2cAA+A6+zC50hk48aNgUDAZrPpun7nnXeaTCaQvPr16999992yLKuqWlZWtmrVKhLbrGr0+v5R4P8IFsMg9Blq+P79+0tLS2H1JtA2YgJmPGZwMV+uYDBs2LDh9OnTHMdJktSvXz94vAghsiz379/f5XJBYXrnnXfKzp+noSBU3GOqHh76xwIaw0NiRKdTp04jRowYP378yZMnCSF2u720tDRRVvb/w28BHMc1adLkwIED+BHInAIXAfXcvHlzWVnZkCFDzGZzOByuRnoCi5nFYklLS7v//vvhufzPSF1gYCzLXn/99d988w3lE8FgEAtEcCDHceFwmGIvtfOnMM8qcxGwa6vNds8998iyXFpaajAYXC7XE088Ub9+feoWQxzhJWU0XE6WZb/88sujR48iAPzGG2+sW7cuIcRgMMBu0K9fP5fLJYoix3HvvfdecVFRtefB/YGAmlmvbNjQZrNJkqQoCsozUIM4JfQkdmGoFRvuQVmSli5dmpOTA4PSww8/HI1GSSwdJy0tDSlpSCemkhT+ChYOjeQ/vvr/EKixmE6kJYEoKIoyadKka6+99tlnn41EIpIkwQWYQozf/0NqwPN8Tk7OuXPn1LiCJUksV4mAZVmfz/fqq6+OGzcOZIpGXlUXQLzz+XwjR44MBAJbtmz5D9wXJPnrum4wGNLT0wsKCqjfCKuLRqO6rjdo0GDPnj00DxGmfrr8qurWVSbEIE8L8/J27doF+34gEJg0aVK7du3wV+wd3A/JT8VoNAqCQHQ9HA6vW7cOZC4QCIwcOdJoNMK3BuaZm5vbtWtXhmEikYjP53v33XfxpyRFh/6sAL+CGosqzs7OhnHPbDaXlZXB5E0IodlYJJaFQOUs2Dq3bduWn58Ph1u7du0yMzMtVis9MpZlBw8ebLfbEb712muvITyJXoOUPUl/FOBiPjOatEzdhzNnzjxz5kxeXh62tLi4OAUq9v+QMqBMxq9UhWVZPnr0qN/vv+eeexDbCsG3uiaJuDVN0+x2u81m69at26ZNm6pr8CSAJdBYUAj9mqZB7QA+C4LQoEGDkydPBgIBxAfDBYIs1BSudpW5iGg0rlq5Mi8vD0EsLMs++OCD9w0ZQgihuT8g7igbkOzdLMvzPGGYffv27dy5k2XZUCjUsGHDjh070lB0eMAIIaNGjQoEAqIout3uN954A7/809Oyi4GL88pGIxEEsHEc53A4VFWljhASizfHt/AbTdOQnmowGJYvXw7PkyiKI0eOhAZJ/Yosy7Zo2fLqq69GnPHx48c/+eQTxLQAzy4noOCPDoi0BquAQRVLdrvds2fP3rRp0+eff04ISU9P/3Nb9n5XoMUqj8FoEw6H4cCr6jg8z2/atKlr164cx6Wnp1dIgP/1wDAMaiPKssyybPfu3Y8dO3bu3LnqGj8R0HS6eP0DOWqEkGAwCBNWTk5Op06d/v73vxuNRqR+ww6GsiBV3Ycqc5HDhw4tX74cNjVRFG+55ZYJEybIkkRiNB3x/nj4krPRdT0aiaxZsyYcDiMrYvjw4YgWgJeYhi3Wa9iwVatWqP508uTJrVu3+v3+v6xJmuM4wjC6rpeXl3OxAHOTyQTbC7YdseR4GFyExn1+/fXXBw8elCTJaDS2bt26Xbt2RqNRkWWGZak2Qwjp06cPYhatVuu7776LOwA7D/erM/9//6DHxRSAQyNmRBTFdu3aDR8+fMKECadOnUI43H91pn8h0DStqKioZs2a+FGPFXSp6jiKonz00Ufdu3eHJA4rcfVGSZjNZuqvvv766xVFOXToUDWOnwiAtHA3IHmTEIJySoimi0ajFotl3LhxX3755aefflohzlCvehZ2lbnIuHHjTp48abVajUaj0+mcO3eu1WaD91/XdYgGNpvtkrmsWKfX6/V6vVu2bEE9rrS0tD59+tBSAQg2vRB0pCijR48uKioKBAJZWVl5eXkIGP8LAj1j1J2FvIOq/jRmlNr0gSKwQyIEvry8fM2aNcjRURSlU6dOTCxjnMRcyiCLffv2RXi4IAjffPNNfn4+iZ2pqqqJalb/aQAiCz7D+oqNRdLZsGHD2rRpM2TIEERq/Xen+tcBnudPnTpVr149JAYiRi4FF+n58+eLi4uvvvpq5G8icqkabRscx0UiES6WLqpp2pVXXnn06NHqGj8J0GDrc+fOUXZLY4gh/RNCrr/++h49ekyaNGnnzp0kjlZT/bsKb0z0B7hVGYZRZYUhjCbJDGHuGzDwXGEBJ/DlPq/Vbntu3tysnGzCEMIySJ5EJDtzOfUNdSLwgs1inT9/PiFEVtVAKHTP4EFMrMYJIQQh7RfAwHfs3KlZyxYWiyUQCJw99fMnH24Le/34I61lAudKkpznkBTVCNEIkSA/6gSJL5ezWRUA7laYklAyJBTLk7oYqPOKGtkvma8g/zJt80L8tKYLvMDoxO/zHT9+3Ov3q7pucdgJIVdddRUN0qdsA3GQDMMQnZSeLwl6fd7Ssi//sT0QCuoMcbicAwfdi+QeludUTTUajZC7CUOisvTohMcDkbAvEhIZblHeQqITKRIl+oWQ9mq0I/8OgSY0UOzC7QLFURRl9uzZgiDMnTuXEIJke1VVqSX6d7s5NK4JU2UYJoUaAfQrtIxK8sTsFAB2CLBtEhOWPR7PoUOHmjRpQl19qbHwEydO1KpVC9USmV/WN6sugGx9ITuNZVu1alWNukg8dsXbD0gsV8ZoNH7//ffNmjVDpiGJ+Zjjk3smT548YcKEKVOmjB071u/3B4NB1Iv0+Xz0WMkvbePx70X1DUJIsuYEyEbhY7z0zVdf/fbbbxlRcDgc586dmzZtGsJ4ysrKUJ8ghb04c+bMxx9/DHJpMBj69++f6ElcyyFDhkx47PGaNWv6vd7Vq1d36XYbIqBBNFGnmka7JlqXJEmnT5+ePXs2uA4qn3NC1ZRZjmEZhikuLi4vL1cU5UJJicQWNrjEoS2yLKvKciAQSJJYR/NAafINLJiqrHA8H4lEbHb7mjVrRFFEZYUb27aB0gaUotLE/1XS1onVahWNxmnTpumx0ofI7yFxxSKBRmysZupNN91Up04dn8/HqfrevXsPHzjQ8KqrIuGw0WxKrfLEHx0Qvw7jgN/vX758efv27dF/gsRSSv1+P3QXKvf9fgB55lAlYT0PhUJWq7WqYjjFW7PZjBoWKEVejWFIHMdBoDEajYWFhW63Gynr33777bPPPisIAqrT07y5KgEKE5CY50BL2rIsNaAp9/jR4XD4fL7qGhzee2rQowSQEIISLJqmffzxx5MmTUquWNx9993XXHPN8uXL27Rp07x584EDB7Zs2ZIG/YMsxJ8px3Hl5eV2u51hGJRCkWU5IRXA9yVJMvACYZidO3fOnz8/GAwaOauu67179x46dCitiBlfpOVyQdcVRdm6dWtZWRlC03r27JmkKxEE6r59+y5+aREqa3366ac/HjtW76qGYCQoaI9MH1QQqnQcsBxFUdavX488I47jgsGg2Vo1F4uuanxc/XyUQkn+FWhp8NPCJehyuRLNMz7qLl4n43i+tKQkrUaNn0+dWr9+vSiKpaWlZrP59ttvJ7FCYdgKUHnQMgggLMuWFhd/9tlnwWBQ59nMzMz+/fvTbCNgPK2QAbtiZmbmnXfe+frrr/MaCYfDb7755vy8PCkQwJ7/prV+f7dAS7MIgpCTk7N27dpBgwbl5OQ0a9bMZDLRHkHl5eW/Q6MrLfgoSdInn3yCQ08hMkcQhJ49exJC0PAK6Erja349eDwe3A5d1/1+f1ZWFn4zZ86cDh061KtXj9YBg/Zc1fHPnz+PglGwUoJdVW+kL82sAmRkZJSVlVXX4FysnSJQMRwOq6pqsViwBEEQ1q5d27hx4+zs7OSinq7r11577eLFi48cOfL5558vWLDg/PnzLVq0aN68+U033YQWivAsoJ8VxAVqVgHlT/gCjuPQ6iToD5SWls6ZMycUCjkcjpAiuVyuZcuWlZWVud1ukrQ8dTJgmOLi4hUrViBSiBDy8MMPJxEH2Fj13MGDBz/77LNOux2BRnMXzIcJklYiSo4KCJwQRREfYJlhWTZRz8FEwMdKz4Jwh8PhSCQSDAQsNmulz9PiZRAiRFEMhUKYdpK3xK/rwiAcn1ajhq5pY8aMgSzpdDoFQWjfvj2J1fmgNdRITHbGFwVBWLlyJa0vctNNN9WvX5/8slcEbiZ89YQQSZIGDx68cuVKORh2u90ff/zxwz/9VKdOHWxX9d66PwTQBqioHaIoSps2bQYNGjR16tT169fzPA8Wglqwv8P9Abm0Wq2hUGjjxo2FhYU48aqqlYFAoGXLljk5ORBxEB1TjSKFy+WCsKUoCpqDGY3GQ4cOrV+/fv369SROGRJFEZGHVZ0/7TlGayBV43lBCo8fEGWzq2t8FEEBWed5HrdVVdVQKITSStOmTfv00095nqdxWYnmiU7njRs3btCgwUMPPXTq1KlPPvlk586db7zxRjgcvvnmm/v06dO8eXNwLOAJOguAzpAkFi1CCG1+MHLkyJ9//hlOV4vdunbtWjQ5SaGBaDx88803P//8M3p1NW/evF79+kkeBqVTFKV///4LFy4Mh8Pp6envv//+8NEPNmzYEEW2CSHI5b5g2a8MbDZbIBAoLS3NzMwsKSmBIGa1WjVSNZOot9zrcrncbvfp06fdbreqqunp6UmCX2kRJyhMsEp7PJ5EuggernAtOY5DOFy/fv2OHTtGYnFEDz74YG5uLi1RRdvdEEJkWUbpTKLrPp9v48aNwABGFO6++24Mq8faWtBNQ7EdQogkSbVr177llls+3rIVhrVXX331ueeeMwpG2nfhLwU4r0AggAIzPM8XFRXNnj379ttvnzlzZl5eHiEEKZkIz/9vz7ciACvOnz+fk5Pz4osvolYYTLIpjEZlI2AC2p9U11QxGrRqRH5OmjTpgQceuOKKK8DL8X9qCeHIdSO/zGyrRi6ix4oc67GmltWoqJFY5QhU6EECLEr82mw2v9/fuXPnJ554Av1FLBYL1LhKxwECoNYU6MYVV1wxdOjQkSNH+ny+kydPfvXVV3l5ecePH7/qqqs6duzYt29ft9sNkoLuW6FQKBkXAX+bOGnygQMHVFm22+0Cx8+cObNu3bogvujnnFrOp8/rXbJkCbqycBw3YsQIQkg0EhFNlWMzNbmkZ2R0795908aNgUAgFAhu3rx5/PjxsKqRmGE6yXvPnDmTlZV10003vfrqqzzPG0QxGAiYTKaqchGe4/0+37fffjtmzBiEi7Asa0ysllFjEQ7AHw47HI4aNWokt8bSFEIY36LR6I7Pts+ZMwcGNFCx1q1bP/roowzH0valNKAe3qYLV10n//M///Pzzz9zHKerWrs2rdu0aQOTFHUwxteOxgcg2YgRI/7x4UeRSMRkMn3wwQdTpkyxOx1/wfIzJEZokKZDCPF4PJmZmZqmPfPMM4899thbb701cOBAUGRY7X9vjATSidvtRtgFzJ4p+LfQpZTiAIILqjHynrZLgihdVFQ0evToxo0bP/TQQ5CvEeOA9HVYZas0fnZ29kcffUQLKle7UwR2gnixrKioKD09vbrGj99/hPjjXceOHZs1a1bLli2HDBlCTyfJucR3IIZhE90HgsGgyWS65pprrrnmmocffjgcDn/++eeffPJJly5dunbt2rFjx549e5rN5vLycovFknDv4HxftWrV+vXrBUGA2WT48OG9evVSVRVt38llN2e9GL755ps9e/ZomhYMBhs1agSDTJIOu3pcT5EBAwZkZ2dDJ33//fePHDlCYio5TExJBCJoDB6Px2yxqKqqa5rRaGR/WZb8ckBVFD3W9FgURTQEVBLHWVP1CBiPCJOff/5ZSQDRaBTFn2VZzs/P37p165NPPtmrV6/x48cXFRXl5+dnZWU5nc6WLVvSHEyqQeu6TtMMSYwVlXs8q1evhjWP5/n77ruPzo3neRpChojt+JlHo9HrrruudevWCMoqLy/fsmULrvGfuBpjIjAYDLgawDGXy1VSUoKaRfPnz585c+b3339PYm68//JcKwM4DlFSk1LPFCq40N4bKBSNeNnqnSe19a9du/aWW2658847n3nmGXrNMeeUO9Dk5uaePn0a2M7+Br2ZEU1D4ppYnzlzJicnp7rGp1W/0GHMYDCUlZVt3Lhx+PDhDRo0mDt3rtVqRRHo5JG7RqORVgDhY70mfT4fWDgfq3gvCELnzp3z8vIOHjxYu3btJ554YsqUKWVlZeALyWK0vvvuuzlz5kDV0HW9SZMm48ePJ4Soqmqz2cDEkH8AW3CVYNGiRWlpaRikX79+JrOZ6HqSNr0QOnRdVyT5uuuuq127dnFxsSAIR48ePXToUKNGjeiyLzkfNdZ0RRRFhmU5llUVpcoxWjxvMBiQcQaWEAwGk/R2pK+GxmDg+aKiotdff/2lpa8kWm80GoWljkYEINgBFsmCgoIRI0Y89NBDDMMYRDE+YgpMCIE3yCPx+/27d+8+ceIExLfs7OyuXbuSGK5TQQbxYyiWhaFKSkpq1KihaVr//v3/+fVXdrs9dP78smXLBt03GHnvVdq0PwHAHYK1I9KJNnZt1arVyJEjJ06cuGbNmlq1atHM3P/2lCsC+m8XFRUhtAmCYFV1JpZlS0pKLBaLzWbDd61WazUaOWVZ/vnnn995550NGzbUq1fvvffea9y4MUIDAoGAw+FAjJkoiqnhYXp6eklJCeJKfiN+X2FLS0pKEpmVUgCEHjAMY7fbDx8+vGHDhj179kSj0Tlz5lx33XUI6wiHw9AzkqyRpiIixAA8FX4KSOSoEMzHmtKKojhmzJj+/fvfd999I0aMWLlypaIoFUknTI2apnlKyx4aNbq8zAOexjBM3osvqgxhY7XHqSH1kiyE+uHB+UVR3Llz5/fff48wg+zs7D59+sixGrQsX7lQwLM8UXWGYQSOVyRp0MB7Txz/we/3u632pYtfHtivv6aouq5zAp/cwsvpRI5GeTTa5DjCEE3TGC4VfVaM0W5oVMnVebgKUcEwEolwJlONGjV8Pl8iMVBVVbPZLEmKzW73eDwQCkSGM3B8eVlZ27ZtH5848YZWrQhDSCyghX4XVQDwQZUVohOOYZcuXarrOisI0Wh0+IMj6RLiX3qx3RYkkmXZ9h1vrp2bW1BQUKNGjbNnz7674Z0+ffuGgyGGY2mSPDgQH+uJ8qcEsFvsG8dx8SVRdV0fN27c0aNHJ06cuHz5cqvVCqqHVp40jpEmavy3QNM0i8WCKwmTTmoR2zQvOt7+Sf9KmRPgxRdfrHQQbAukllAoVFZWVlJS4vf7Dx06ZDKZbr311rfeeqtFixa0hgWK/ZA4+pMan65Vq1bNmjW/+uqrzp07I2xSEIRqFIyoeRnRzxzHffzxxy+//HKi5xVFMZlMNpvt73//O4JoESCTSMMLBAJlZWXHjh3Lz8/neb5bt26PPfbYrbfeeiGRQFU5jqOZ6kmcDlRehKu8AgXATOiPtDFSVlbWli1bRowYMWHChMWLF/PxG0fDN1VVHTt27IkTJ7KystDJZOfOnRkZGZeUtS8GSZKArzTelBCyevVqzNtisUSj0fvvv9/j8SBuJBEXETke2yHLstPpLC0tPXsuPy0tjdFJcXHxhg0b+g8cGPD7rYINRO33RshwWjBl2mw2XVWLioosJjNvqHxLNVUv93lr1qx5/vx5m8Pu9XqNZlMoFGratOnjEydcffXVoslECNF13ev1JolxMJpMqqKcOHHiwIEDPM/D/vDOO++8+97GKs3fJBqLi4t1XT98+HCzZs2WLl3apUsXu8MBNka15j83C0kCqBHrdDrz8vJuvvnmZcuWPfTQQzQXh8SMMIFAwOVy/QWjEk6fPl3p78FfYe81m83p6emNGjWyWq21a9feuXPnlClTUHQVpn/qzK8W6NChw9atW8FFzGYzQlKra3DI4jh0QRC++OILm83WqFGjRM9zHAdvtt/vBwUGrUgkcLjd7rp163bt2nXXrl3btm177LHHQGYpG6D6R6LmWqkBTWfmef6JJ5544IEHdu/ezcMgSCvNwYqyZMmSf/7zn3a7Hbk5CxcurFGjBi8IetXFduq3oUzv8OHD27dvBwvx+/2hUOjrr7/Gjlut1mA4VOk4jKrRZAiz2YxAN6PRKEWiPr9v5cqV/QcOpK8Ih8O/t6IUsB0pihIKhaiP58Ybb8ypVbPS54FDTqcTUtKiRYu8Xq8oinv37hUEwZmWRghRFUVnLqELRiMRURTfeustJGM6nc7z588fO3ZMryKpj4TCTqdTkqS6detGo9Effvhh7969HW+5BX9Fc2wUdEtZtv1DA8uyOAhRFNesWXPnnXc2b968Y8eOVEnFVbLb7aii+l+e7n8cnn/++Up/Dx2OxtpCImYY5pZbbvnhhx/mzJkzZ84cQgjCQauXi9xxxx2PP/641+tF9Gm1J/cgLBNRTN98803dunWRoVIpoKhd3bp1J02aBCEjuaWR1l3t0qXLiRMnJk+ePG/ePIfDwcbqwNMPMJZUFyOJr4bZokWLm2++ecWKFTzP82D1sIaHw+EvvvhiyZIlsGKpqtq1a9c7uncnhGiJw2eTAL4ClQpIsGLFClmWia6fP3/eYrEgns9ut/v9/iQOOkXTMCU21hGTNxgKiopcNrvFYjl06NDnO3bc3LEjHv4dlpsFC0RYXjQa5VnWZrP17Nmz/6B7K30+4PdDnIHdTNG0V155xWKxyOHIyJEjP/744/SsLI7noQoksXsKgnDu3LnNmzfDgoxcAbPZnIhbJwK32x0MBhE5AwvvO++80759e+iOQNlIJGI2m2kVr6pt0B8cKLfgeb5JkyZPPPHE2LFjP/jgA5pIiz7VgiB4PJ7fobPkt4ZE+MD+slFrfKb3rFmzunXr1qpVqzvvvBPhoNWYj6IoStu2bVVV3bp168CBA8HGqhFpGYZBkB4hJBgMrl27dv78+fHCdAVAayWv1wssQjUdeECTvAIG5Llz5/bu3fuLL77o1asX6gigAgr9bpLMh6oCDEKw16mq2rdv39GjR/PRaBSsGCykoKBg5syZeGUkErnhhhumTZumKgrH8yzHKbKcyAKTCOA/p3GuP/300/bt2wVB0FVN1/WRI0fquh4MBtEdOkkOedAfoNVNmFhDrvz8/I8+2OpwOSVJev311zt06MBwLILff28VZ5HJgWQOQRBUWZYkKYkGbbXZwqGQyWwOh0IGg2Hs2LF79+79xz/+USsru7y8/Lnnnlvw4oskZidMUjuA5bjXXnsNVbnsdnvDhg3btm2r67qiVTHdV9ORaLJixYq0tLRwOPzPf/7zxIkT9RrUh/wIUYBGpFRt8D8+mEwm+P9QTeD+++//7rvvJk6c+Pbbb+MB+MMEQaje7LM/CiSRqWn2OE39A31s3LjxggULJkyY0KVLF5PJhPoU1ZVyARV/xIgRS5YsGThwIKXI1TI4wGw2Y8IbN27UNK1Tp07o1pPkK8FgkBaSQGJZon2LRqPYE5Zlc3JyXnjhheHDh9epU+e6665DGiYeg3xZ7RYtnBH12/NgIVhwKBQaN27cuXPnkFldq1at+fPnO5xOQkg4FIqvk3j5QCk+HP0bN248d+4csi579eo1ZswYRVFsdrumqgjbVhLkkPOxtlwwm6B5SWlp6ZfbdyBm46uvvjpw4MC1zZtFIpHfp90ZyuAFRqjr4XA4EAjoCfJUVFXViE4YYrKYCSGcwE+aMvno8WNFZ89ZbNYtW7Y0bdbsnnvuMZlMyQNjfF7vypUrIdQIgjBp0qSWLVuqqlpVaUBTVIMoypJEUxdZln311VcXvvQiiTXjpOVL/2qKCCEESRjnz59PT083m80Gg2H+/Pl9+vSZOXPmU089BVoJURpS6n97vv9pSIQSgUAANSkq0BZBEILBYKdOnXbv3j1s2LDXX3+9emvtiKLo8XgefPDB9evXP//88xMmTDCZTPG+218J4B+CIBw8eHDRokV///vfacmWRKAoisVi4WPFdRCkm4ixYZ40Mbxp06a33XbbwoUL58+fn5GRgaoKCOfheb4abTNms5nWPZIkyel0RqNRFsneMA5Omzbtu+++Q4F3g8GwePHiOnXqaKqqyLLJbDaaTCnYc+HWR1h3eXn55s2baRh13759EdFMYs5JEsshuhhoqww2Vg2X5/nMzMzbbrsNxcwVRXn77bexpylHkf+mQF0jaGkA7w6TAIBSNCGjrKysefPmQ4cOrVGjhsVikSTpxRdfPHPmjNfrtVqtSWJ+Vq1a5fV6Ido0btz4xjZtBIPBaDJVNT/GEKsu069fP9TV0XX9nXfe+fnnn0mscB4hBELQX1AXIYSgfgFKUCCAcsGCBRs3bnzvvffoTUbZVPTZ/UuBlgDQY4IGsqM+EL6C3hOPPfbY8ePHly1bdnEE0a8El8vFcdz06dNfe+21PXv2kGrFWwSghsPhBQsWtGnT5o477rhkXqSiKLhWsiyD0oLuVQqgDKDbWMvChQuPHz++evVqtCWladrVuy4KSFqADYlF3I7D4Zg3b95HH31kMplEURQEYerUqS2vvx5xY5jx+eLiJLnZSUAQBFQl2b59++nTp00mk6ZpWVlZLVq0IAxjdzhCwSBhGIfTmWSjA4EA/opaAihdRQgZPHiw2WxGAai1a9fC6Pw77AcOOo4IdzAS2tKgUsCfeJ4vLCzUNM3tdkcikUcfffTaa69FEWKfz/fkk0+CDSeRNZYvX26321GA9oEHHiAp45OuQ19s2rRphw4dUJGNZdl169ZBNqHZSUnU8D89IKwD3ixVVf/2t7+NGzdu7ty5+/fvxwMWi8XhcPw1dZFKgVqxEMsO8Rm9wXER7HZ7Xl7e0qVLv/7669+CGt588819+/Z96qmnioqKqjEkB9ExS5Ys+emnn6ZOnYqSQskZCSXKUETASBLtG8/zKGyIMdEc9pVXXnnrrbdQGwnXE5SheqVqo9EI567NZtuxY0fdunVZORxhVO2Lz7a/ufw1njBKJMqo2q0dbr5/+DDCENFkJCxjNJsIQ9IzMwhDiF75P4Tr+f1+Qkh8FIoiyUQnuqpFQuFVK1byLMfoxGaxjhz1IMOxhCGSLPEGQSc6ulwkmrrdbseOoJMHjYa+rtUNTZs1Kykrg79kxZtvSZGopqjxodbIBkdCDWi30WhUY2I+GytcWCXQNY1hWTSORfF2SZKS0GiIWoQQVFqm3CIRlgAIIVlZWfgA99JLr7zsrlHDHwzyPP/tt98ufmkRQxii6YQQlHek6oskScuWLYsEgqosR8PhJk2atG7dWlJk7HlV1xtW5agqE55VNHXUQ6N9Pp+iKBaT6c3lr6myoqsadoDlOMSfJRqnggWcxDrJV1g41cmqOs9EAFsiFeKg7SXRrSORCPyH1EB3yflQnKTnRQgZOnRo9+7dx4wZ4/V6aRmbFCwM0WiUbhSNa0ptf7APdGnInNWrCCSWGIuwETo4qmhQ7KV280STudjSxfM8iCm+azab27Vrd88990ycOJF6UCBBxisudEvpb5I3UsN99Hq9iqLMnDnziiuu6N27N76rKArGp0uTZRnlVuOLNeAI8KQaa1MNAF6tW7du2bJlL7zwQp06dRC8m7yxG6gzhgWZSk79wfNoZ3FCSIsWLYYNG/bUU08VFxdT3Pb5fIjmiJ8hIv4xvq7r8UWdo9Eo5qn9sm2JLMvQpDFVvHf79u09e/ZkBaNREITx48erqur1eo1GY7169d5YsSLJ7CsF8D1Ey1FtAJNTFYVh2X/961///ve/kekty3KvXr1oND3Lsqi2ltqtGD16tNvtLiwstNvtGzdu9Hq9SG+k+8LG9fIDt6sg1KQg4zCxuGmj0WgymdBPIknuPTKP6OtSpo8Oh2PSpEnofxmJRNatW/fF558jjwcl1eiN1TRt9erV4XAYtsv777/fbDbjXFLwIvI8bxSNLMOaTKbmzZv/7W9/Q/aDJElvv/02w7LIHQuFQzrRk3ApasTDj/FU6T+QaHL5r6guNoaw+szMzLlz5yLqBi69qo5TXZsDczlKGJBYLksKJdFAN2n/R/p79pf11ePPt6pAGQzLsqNGjapTp87YsWPBO9FJCclJFSg44JLHB0rtcFwoBzdz5swmTZp06tSpvLyc53m4G9GBAoYm3HEmVgORcnREVZC4OwWPyKJFiyZMmPDmm2+2bNny3LlzmHDyxOSLJ5zCoY8cOTI9PX3q1Kk8z4OR2O12kH6WZaEEr1ixokePHggYA0HWNK28vJzEKCHtVKTFEtqhS1ksFshJ5eXloih+8sknR44c6d+/P3u+sLBHjx5FRUU8z6enp/M8v2rVKqJpaiJQlEr/gdliJZj0hYppoojGG0uXLv3hhx9cLpeiKA888AAS5Sjhw+1KAduikUinzp3T09MRSXnixIl169aZzGZgM9ALlAt8BXmqFUwuqb1XlqRwOAz+B8YuJfZPAEWAfxVkuioBx/N39+49YMAAr9frdDo9Hs/zzz9fXFQUDAaRvk6Xtm7dOoi9sizXq1evV69eJIamKXi/4xHa6XQOGDCA53nwztWrV3vLy6Gwa5rGECaJrA12XoGLEEKgGVxcT6yq87zkKuIXkmQf4n1U9JepSBsMwzDM7Nmzt23b9vrrr5OYxTyFceI/pxy/gFuALnXU4gGxpkoAT1h8tVBd10OhEP/LtFOoTalNFUOh46rb7X744Yd37NjxwQcfsCwbDAbtdnswGEQdlIulIkwyyS5RV6vH42EYJjMzc968eU2aNOnWrduOHTsIIajZEQwGESuFm0u1ZOhhINNutxu3LBwOy7J88uTJUaNGbdy4cdu2bTfeeGNBQUFOTo6u6yB9Sc6FxHgwvRQpnLLZbH7qqaf27Nnz/vvvcxwXCoWQFY9YHjQYHjJkSNOmTTt16nTkyBH0yBAEAUmLDMOIMYoNaYMST1BOvMXpdB45cuSBBx6YMWOGruv8/PnzS0tLgRCSJDVo0ODxxx83Go1KAtN2IpyQZTkajTqdTl3XcfCCIPTq1at///4Gg+HwoUN79uzJyMhA0Yh7770XCIdSVHSQJBGriQDP33vvvYsXL/Z6vTVr1ly3bt0jY8fSSvJU8AeRYgmLsKhfKdyB5wHYWO+T5DI+bhQKqVKWltrbZ8yYgczN4uLi77//funSpdNmTFdVNRQK2Ww2zOqzzz4rKyszGAwlJSWjRo0ymkyoVJYkaD0JMIQJhoIWswUL6dOnz5IlS0qKizmOO3HixNatWwcOuleSJKvFSgjh2IRc5GJBFdeSlmuDxQa4W73+LdC++JuZ5JZSw1r8bLVYytXlA8w7V1111dy5c0ePHn3zzTcXFxcnyT5LMp8kc7t8wBVzOp3AE1SN5OP6rV0mgHNQWx9+Gb858SpmClNFclVpaWlaWhrE17Zt244YMWL69Ont2rVzOBxlZWUul4thGBr+G6/iMzEbaaLxqR3bZrOh8p7FYlm4cOHSpUtHjx7dvn37yZMn169f32w20xhImqBN4jqV+Xw+QRBMJhOu/5o1axYvXpybm7t582bYhbKzs5GnlTxWSv9lSC4wLQUuEo1GmzVr9sgjj8ycOfOqq65q1KgR9Am0MyeEwIU8b968RYsW9ezZ89577x0xYkTt2rXRbwmUASX4qMiLmbNx/s6dO3fOmjVr7NixnTp1slgs/Icffnjq1KmaNWsKglBSUuLxeH788UdN00iCBSTiIijLiNpBMCCazeb27duHw2GT2fziiy+GQqGsrKyzZ88OGzYsLS1NUmR6E0KhENzOKYgtLMdpqjpkyJCFCxcituHs2bMb1q/v278fqAYeA8lmGEbgBVWvGIqawmmxLIvXwa91SawlsX7I+K7OsiRxEP0lwWg0vvHGG126dEETixUrVjS7rnnPnj1hUYxEIvv379+/f7/H43Fb7RkZGXfddRctdgk1NoXEN3pFVVU1WyxdunR5dckSq9XK8/x7773Xb0B/k8mkE6LpGrlUH3s91rKXibWgl0Jh+hsQoOR+ixQA76og1Cd6GNemUrWpSsCyLOwYN9xww7hx4+666y6Hw5FCvgi9w9gioJxW9aBqENxAIIBQRooGVTVyapoGCw96Q/j9fhQpoQ/EC9QpWFDBQlwuF+wqsLGMGDHi8OHDo0aNevvtt/H2Cu+iEO9ZrBSQEogUH6fTiQ+apo0aNWrAgAFjx4695ZZb7r777t69e7dq1YrE8t4IIegBQVcKe9GPP/741VdfvfDCCzzP5+XltWvXDq8OBALoP0i3PVFQPix11JmEX6bAfdGweejQofv27Rs7duymTZusVivLsqjvoigKLBayLI8dO7ZHjx7Tp0/v27cvOPSVV16JN/r9flASjuPi3XjRaPTYsWNLly7ds2fPhAkTunfvDg8WL0kSFC6EQ5SXl4PYGRIcvJrgIgkcZzWbYbKxWSxQAsLBoMlsPnb06K5du5DVmJubO3ToUEKIyWSCiIFEPHD4FGRPTVVZjuN1vW/fvuvWrQuFQhaLZcOGDbd3vwPN5rRYI1iGYcxms5k3SErFioEpeDu1mNGPxIpfkguCZ+VDUbkbkosmCHrS2KokEAoGzRZLkyZNnn766Zdeesnlcp05c2bWrFnXXXddrVq1kNr25ptver3eWrVqaRHpzl69atWqhe/SKPIqr1fXzCYziV0nVVaGDRv24QcfhMNhE8cdOHBg586dHTt2lBVZ4AVZU5OoIyQmaWIHEFavC0o8F8HOVKMuQjkW+8tk6URwMctJ2b4PPDEajSNHjvziiy+OHz+e3MuaaD50GvG+66qOA4KOQXbv3u1yuYLBoNfrrWq5WTTKZVk2HA63aNHCbDbHswqQfioipHaOGJ+SP5gxFi5c2Lp16xUrVjzwwAO6rpeUlKSnp9Neh5evi0DJcLvdKEMHAzvS41VVXbly5cmTJ1esWDFlypRQKHTNNdd07dq1Vq1atWvXttlsDMMUFhbm5+fn5+fv3r370KFDxcXFNWvWnDNnzh133IFUKlTjBtZR+pYkBowKB+RXYBqJdaUjhEyfPv3uu+9+8803Bw8e7Ha7HQ5HYWFhVlYWwqscDoemaQ0bNkQx4GXLlvXq1Ss7O7tVq1bNmzdv1KgRSkOi33ZJSclPP/303XffHT58+PTp04MGDZozZ47NZoPicu7cOT4nJwfhKAzDoCtWZmZmfn5+IkdQIi5CNI3neeyg2WxGSInNZlMVZefOnS6Xq0aNGrIs33DDDY0bNw4EAjaHXY8VBaG6Xgq8l+U4QkggEBg9evSXX36JtZSVlZWWllKjHokVp3S73bxGwtFIBeEoFV2E44wmk8vlojk+drudTcwVgCJIXXY4HJqiKKk2mKOoNmLEiJ07d/7www/16tWTFHnVqlUPPfSQw+EoKCj48ccfr7zyypKSEqfT2a1bN03TWI5TFQUVU1LzpvI8K8kSTB+RSCT3iivatm27b9++YDisquru3btvuukmVVUFPlnXMhquTZ+BYdNls1fgIoqiZGZmprA/lUK83ZlcnhZYwRST8t1GVi/2/OWXX77rrrtgg04BqI0oZS4CbyXP86WlpdOmTUOuFYmdy+WDqqroc2WxWJ588slmzZqhUwXoBrgIijXQkr1VAsRMB4NBh8OBZEDQZWqNufXWW7Ozs9PT01F9I167JZfhXUfrBFVVHQ4HOlfiW5FIxOVyRaPR+vXrP/vss6WlpadPn/72228//PDD0tLSgoKCsrIyCNwZGRkul+vaa68dP3781VdfXbt2bUKI3+8HH0L8AmUb8OUkyQ6mJxv/YwqA0wT1y8vLe+ihh3r16gXrZVZWFh7AYinRa9y48bJly/x+/+eff/79999/9tlnL730ksfjgfTD87zL5apfv36TJk0GDhzYuXNnZDuFw2FI0jk5Of8LKlHZa4GgOD0AAAAASUVORK5CYII="

CO = {'addr':'Zollergasse 9/26, 1070, Vienna - Austria','email':'viviane.zahra@zahra-interiors.com','phone':'+ 43 (0) 67682337441','vat':'ATU77022918','eori':'ATEOS1000136890','bank':'Raiffeisen Bank /  Vienna - Austria','iban':'AT47 3200 0000 1348 4274','bic':'RLNWATWW','owner':'Viviane Zahra','name':'Zahra Interiors'}
C0 = {'d':'404040','m':'606060','l':'999999','v':'CCCCCC','bg':'F7F7F7','w':'FFFFFF','h':'4A4A4A'}
TN = Side(style='thin', color=C0['v'])
MD = Side(style='medium', color=C0['d'])
BT = Border(top=TN, bottom=TN, left=TN, right=TN)
BH = Border(top=MD, bottom=MD, left=TN, right=TN)
BM = Border(top=MD)

def F(sz=9,b=False,c='d',n='Calibri Light'):
    return Font(name=n,size=sz,bold=b,color=C0.get(c,c))

def gen(data):
    wb=Workbook();ws=wb.active
    dt=data.get('docType','Invoice');ws.title=dt
    cur=data.get('currency','€')
    cf='#,##0.00 "€"' if cur=='€' else '"$" #,##0.00'
    ws.page_setup.paperSize=ws.PAPERSIZE_A4;ws.page_setup.orientation='portrait'
    ws.page_setup.fitToWidth=1;ws.page_setup.fitToHeight=1
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_margins.left=0.55;ws.page_margins.right=0.55;ws.page_margins.top=0.4;ws.page_margins.bottom=0.4
    for i,w in enumerate([15,36,4,4,8,15,5,15]):
        ws.column_dimensions[get_column_letter(i+1)].width=w

    def W(row,col,val,font=None,fill=None,align=None,border=None,nf=None,mg=None):
        c=ws.cell(row=row,column=col,value=val)
        if font:c.font=font
        if fill:c.fill=fill
        if align:c.alignment=align
        if border:c.border=border
        if nf:c.number_format=nf
        if mg:ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=mg)

    # ── LOGO top-right ────────────────────────────────────────────────────────
    try:
        logo_bytes = base64.b64decode(LOGO_B64)
        logo_img = XLImage(io.BytesIO(logo_bytes))
        logo_img.width = 320
        logo_img.height = 23
        ws.add_image(logo_img, 'F2')
    except Exception:
        pass  # never crash if logo fails

    r=2
    W(r,1,dt,F(18,True));ws.row_dimensions[r].height=28
    r=3
    for c in range(1,9):W(r,c,None,border=BM)
    ws.row_dimensions[r].height=4

    r=4
    info=[]
    info.append((dt+' Nr.' if dt=='Invoice' else 'Quotation #',data.get('docNumber','')))
    info.append(('Date',data.get('date','')))
    info.append(('VAT Nr.',CO['vat']))
    if data.get('clientVat'):info.append(('Client VAT Nr.',data['clientVat']))
    info.append(('EORI #',CO['eori']))
    if data.get('workStart'):info.append(('Work Starting Date',data['workStart']))
    if data.get('workEnd'):info.append(('Work Compl. Date',data['workEnd']))
    for i,(l,v) in enumerate(info):
        W(r+i,6,l,F(8,False,'m'),align=Alignment(horizontal='right'))
        W(r+i,8,v,F(8,False,'d'))
        ws.row_dimensions[r+i].height=13

    W(r,1,'Client:',F(9,False,'m'))
    W(r+1,1,data.get('clientName',''),F(11,True,'d','Arial'),mg=4)
    ws.row_dimensions[r+1].height=18
    addr=[l for l in(data.get('clientAddress','')or'').replace('\n','\n').split('\n')if l.strip()]
    for i,line in enumerate(addr):
        W(r+2+i,1,line,F(9,False,'l'),mg=4)
        ws.row_dimensions[r+2+i].height=13

    r=max(r+2+len(addr),r+len(info))+1
    W(r,1,'Project:',F(10,False,'m'))
    r+=1
    W(r,1,data.get('projectName',''),F(11,True,'d'),mg=5)
    ws.row_dimensions[r].height=18
    if data.get('location'):
        r+=1;W(r,1,'Location: '+data['location'],F(9,False,'m'),mg=8)
    r+=2

    items=data.get('items',[]);scope=[l for l in(data.get('scopeLines')or[])if isinstance(l,str)]

    # ── LINE ITEMS (Invoice AND Quotation) ────────────────────────────────────
    if items and len(items)>0:
        hf=F(9,True,'w');hfi=PatternFill('solid',fgColor=C0['h'])
        ha=Alignment(horizontal='center',vertical='center')
        for c,h in enumerate(['Item Number','Item Description','','','Qty','Unit Price','','Total'],1):
            al=ha if c>=5 else Alignment(horizontal='left',vertical='center')
            W(r,c,h,hf,hfi,al,BH)
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
        ws.row_dimensions[r].height=22;r+=1
        af=PatternFill('solid',fgColor=C0['bg']);wf=PatternFill('solid',fgColor=C0['w'])
        ra=Alignment(horizontal='right');ca=Alignment(horizontal='center')
        for idx,item in enumerate(items):
            fl=af if idx%2==1 else wf
            W(r,1,item.get('itemNumber',''),F(9,False,'d'),fl,border=BT)
            W(r,2,item.get('description',''),F(9,False,'d'),fl,border=BT,mg=4)
            for cc in range(3,5):W(r,cc,None,fill=fl,border=BT)
            W(r,5,item.get('qty',0),F(9,False,'d'),fl,ca,BT)
            W(r,6,item.get('price',0),F(9,False,'d'),fl,ra,BT,cf)
            W(r,7,None,fill=fl,border=BT)
            tv=(item.get('qty',0)or 0)*(item.get('price',0)or 0)
            W(r,8,tv,F(9,False,'d'),fl,ra,BT,cf)
            ws.row_dimensions[r].height=18;r+=1
        r+=1

    # ── SCOPE OF WORK (always shown if present, for both doc types) ───────────
    if scope and len(scope)>0:
        W(r,1,'Scope Of Work',F(10,True,'d'),mg=8);r+=1
        for c in range(1,9):W(r,c,None,border=Border(top=TN))
        ws.row_dimensions[r].height=4;r+=1
        for line in scope:
            ip=line.strip().lower().startswith('phase')or line.strip().startswith('##')
            cl=line.replace('##','').strip()
            if ip:r+=1;W(r,1,cl,F(10,True,'d'),mg=8)
            elif cl:W(r,1,cl,F(9,False,'m'),mg=8)
            ws.row_dimensions[r].height=15;r+=1
        r+=1

    # ── TOTALS ────────────────────────────────────────────────────────────────
    te=data.get('totalExclVat',0)or 0;vp=data.get('vatPercent',0)or 0
    va=te*vp/100;gr=te+va
    for c in range(5,9):W(r,c,None,border=BM)
    ws.row_dimensions[r].height=4;r+=1
    W(r,5,'Total excl. VAT',F(10,False,'d'),mg=7)
    W(r,8,te,F(10,True,'d'),align=Alignment(horizontal='right'),nf=cf)
    ws.row_dimensions[r].height=20;r+=1
    vl=f'VAT ({vp}%)' if vp>0 else 'VAT reverse charge'
    W(r,5,vl,F(9,False,'m'),mg=7)
    W(r,8,va,F(9,False,'m'),align=Alignment(horizontal='right'),nf=cf)
    ws.row_dimensions[r].height=18;r+=1
    gf=F(11,True,'d')
    for c in range(5,9):W(r,c,None,border=Border(top=MD))
    W(r,5,'Total Amount',gf,align=Alignment(vertical='center'),mg=7,border=Border(top=MD,bottom=MD))
    W(r,8,gr,gf,align=Alignment(horizontal='right',vertical='center'),nf=cf,border=Border(top=MD,bottom=MD))
    ws.row_dimensions[r].height=24;r+=2

    # ── PAYMENT TERMS ─────────────────────────────────────────────────────────
    pt=data.get('paymentTerms',[])
    if pt and len(pt)>0:
        W(r,1,'Payments As Follows:',F(10,False,'d'),mg=5);r+=1
        for t in pt:
            if isinstance(t,str)and t.strip():
                W(r,1,'  '+t.strip(),F(9,False,'m'),mg=5)
                ws.row_dimensions[r].height=14;r+=1
        r+=1

    # ── ADDITIONAL COMMENTS ───────────────────────────────────────────────────
    ac=data.get('additionalComments','')
    if ac and ac.strip():
        W(r,1,'Additional Comments:',F(10,False,'d'),mg=8);r+=1
        for line in ac.split('\n'):
            if line.strip():
                W(r,1,'  '+line.strip(),F(9,False,'m'),mg=8)
                ws.row_dimensions[r].height=14;r+=1
        r+=1

    # ── SIGNATURE BLOCK ───────────────────────────────────────────────────────
    r+=1
    sb=Border(bottom=Side(style='thin',color=C0['d']))
    for c in range(1,4):W(r,c,None,border=sb)
    for c in range(5,9):W(r,c,None,border=sb)
    ws.row_dimensions[r].height=28;r+=1
    W(r,1,CO['owner'],F(9,False,'d'),mg=3)
    W(r,5,'Client Date & Signature',F(9,False,'l'),mg=8)
    r+=1;W(r,1,CO['name'],F(8,False,'l'),mg=3)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    fs=max(r+3,52)
    for c in range(1,9):W(fs,c,None,border=Border(top=Side(style='thin',color=C0['v'])))
    ff=F(7.5,False,'l');fa=Alignment(horizontal='center',vertical='center')
    fl=[CO['addr'],f"e-mail: {CO['email']}        Tel. {CO['phone']}",f"VAT# {CO['vat']}                EORI# {CO['eori']}",f"{CO['bank']}      IBAN: {CO['iban']}      BIC: {CO['bic']}"]
    for i,line in enumerate(fl):
        row=fs+1+i;W(row,1,line,ff,align=fa,mg=8);ws.row_dimensions[row].height=11
    ws.print_area=f'A1:H{fs+len(fl)+1}'

    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return buf.getvalue()

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        body=self.rfile.read(int(self.headers.get('Content-Length',0)))
        data=json.loads(body)
        xls=gen(data)
        fn=data.get('docNumber','doc')+'.xlsx'
        self.send_response(200)
        self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition',f'attachment; filename="{fn}"')
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers','Content-Type')
        self.end_headers()
        self.wfile.write(xls)
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers','Content-Type')
        self.end_headers()
